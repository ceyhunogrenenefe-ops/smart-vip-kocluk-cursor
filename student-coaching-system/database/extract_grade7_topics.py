"""Extract topic lists from Kopyası - 7. SINIF_ (1).xlsx KONU TAKİBİ sheets (column D)."""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "Kopyası - 7. SINIF_ (1).xlsx"
OUT = Path(__file__).resolve().parent / "grade7_topics_from_xlsx.json"

# Sheet indices for * KONU TAKİBİ (verified from workbook order)
SHEET_INDICES = {
    "TÜRKÇE": 5,
    "MATEMATİK": 6,
    "FEN BİLİMLERİ": 7,
    "SOSYAL BİLİMLER": 8,
    "DİN KÜLTÜRÜ": 9,
    "İNGİLİZCE": 10,
}


def normalize_cell(s: str) -> str:
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    return s


def extract_topics(ws) -> list[str]:
    topics: list[str] = []
    for row in ws.iter_rows(values_only=True):
        if len(row) < 4:
            continue
        cell = row[3]
        if cell is None:
            continue
        text = normalize_cell(str(cell))
        if not text:
            continue
        low = text.casefold()
        if "konuları" in low or "konulari" in low:
            continue
        if text in topics:
            continue
        topics.append(text)
    return topics


def main() -> None:
    if not XLSX.is_file():
        print(f"Missing: {XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    result: dict[str, list[str]] = {}
    try:
        names = wb.sheetnames
        for subject, idx in SHEET_INDICES.items():
            if idx >= len(names):
                print(f"Skip {subject}: sheet index {idx}", file=sys.stderr)
                continue
            ws = wb[names[idx]]
            result[subject] = extract_topics(ws)
    finally:
        wb.close()

    OUT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} ({len(result)} subjects)")


if __name__ == "__main__":
    main()
